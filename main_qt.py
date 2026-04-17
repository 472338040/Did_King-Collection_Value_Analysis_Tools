import sys
import os
import math
import io
from math import gcd
from itertools import product
from contextlib import redirect_stdout

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QTextEdit,
    QGroupBox, QTabWidget, QFileDialog, QMessageBox, QSpinBox,
    QDoubleSpinBox, QCheckBox, QSplitter, QScrollArea, QFrame,
    QComboBox, QStatusBar, QToolTip, QSizePolicy
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QPalette, QIcon, QTextCursor

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# ============================================================
# 原有核心逻辑 (保持不变)
# ============================================================

def load_grid_stats(xlsx_path="Data.xlsx", weighted=False, verbose=True):
    if load_workbook is None:
        raise RuntimeError("未安装 openpyxl，请先执行: pip install openpyxl")
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"找不到文件: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    col_to_grids = {}
    for cell in ws[1]:
        val = cell.value
        if isinstance(val, str) and "*" in val:
            left = val.split("*", 1)[0].strip()
            try:
                col_to_grids[cell.column] = int(left)
            except ValueError:
                continue
    if not col_to_grids:
        raise ValueError("表头解析失败，未找到形如 'X*Y物品均价' 的列")

    raw_samples = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        quality = row[0].value
        if not quality:
            continue
        samples = []
        for cell in row[1:]:
            grids = col_to_grids.get(cell.column)
            if grids is None:
                continue
            price = cell.value
            if not isinstance(price, (int, float)) or price <= 0:
                continue
            samples.append((price, grids))
        raw_samples[quality] = samples

    raw_samples["绿白"] = raw_samples.get("绿色", []) + raw_samples.get("白色", [])

    results = {}
    if verbose:
        mode_name = "加权平均" if weighted else "简单平均"
        print(f"========== 从 {xlsx_path} 读取统计 ({mode_name}) ==========")
        print(f"  {'品质':<4}  {'样本数':>5}  {'每格价值':>12}  {'平均格数':>10}")
        print("  " + "-" * 40)

    for quality, samples in raw_samples.items():
        if not samples:
            results[quality] = {"per_grid": 0, "avg_grids": 0.0, "samples": 0}
            if verbose:
                print(f"  [警告] {quality}: 没有有效数据")
            continue
        per_grid_prices = [p / g for p, g in samples]
        total_price = sum(p for p, _ in samples)
        total_grids_sum = sum(g for _, g in samples)
        avg_price_per_grid = (total_price / total_grids_sum) if weighted \
            else sum(per_grid_prices) / len(per_grid_prices)
        avg_grids_per_item = total_grids_sum / len(samples)
        results[quality] = {
            "per_grid": int(round(avg_price_per_grid)),
            "avg_grids": round(avg_grids_per_item, 4),
            "samples": len(samples),
        }
        if verbose:
            print(f"  {quality:<4}  {len(samples):>5d}  "
                  f"{results[quality]['per_grid']:>12,}  "
                  f"{results[quality]['avg_grids']:>10.3f}")

    if verbose:
        print("=" * 56)
        print()
    return results


def load_grid_values(xlsx_path="Data.xlsx", weighted=False, verbose=True):
    stats = load_grid_stats(xlsx_path, weighted, verbose)
    return {q: s["per_grid"] for q, s in stats.items()}


def estimate_grid_candidates(avg, max_count, max_grids=200, error=0.05,
                             known_grids=None):
    if known_grids is not None:
        g = known_grids
        if avg is None:
            return [(c, g) for c in range(1, min(max_count, g) + 1)]
        avg_f = float(avg)
        if avg_f <= 0:
            return []
        candidates = []
        for count in range(1, min(max_count, g) + 1):
            if abs(g / count - avg_f) <= error + 1e-9:
                candidates.append((count, g))
        return candidates

    if avg is None:
        return []
    avg = float(avg)
    if avg <= 0:
        return []

    candidates = []
    for count in range(1, max_count + 1):
        min_g = (avg - error) * count
        max_g = (avg + error) * count
        g_low = max(0, math.ceil(min_g - 1e-9))
        g_high = min(max_grids, math.floor(max_g + 1e-9))
        for g in range(g_low, g_high + 1):
            if count == 0:
                continue
            if g < count:
                continue
            if abs(g / count - avg) <= error + 1e-9:
                candidates.append((count, g))
    return candidates


def print_candidates_table(name, label, candidates, max_show=50):
    print(f"\n[{name}]  {label}  候选 {len(candidates)} 个")
    if not candidates:
        print("  (无)")
        return
    shown = candidates[:max_show]
    line = "  ".join(f"({c},{g})" for c, g in shown)
    print(f"  (count, grids): {line}"
          + (f"  ... 还有 {len(candidates) - max_show} 个"
             if len(candidates) > max_show else ""))


def _triangulate(name, avg, count, grids, error=0.05, max_count=None, max_grids=200):
    if count is None and grids is None and avg is None:
        return None

    if count is not None and count == 0:
        if grids is not None and grids != 0:
            print(f"[警告] {name}: count=0 但 grids={grids}, 矛盾")
            return []
        return [(0, 0)]
    if grids is not None and grids == 0:
        if count is not None and count != 0:
            print(f"[警告] {name}: grids=0 但 count={count}, 矛盾")
            return []
        return [(0, 0)]

    if count is not None and grids is not None and avg is not None:
        if count <= 0 or grids < count:
            print(f"[警告] {name}: count={count}, grids={grids} 违反物理约束 count<=grids")
            return []
        if not (abs(grids / count - float(avg)) <= error + 1e-9):
            print(f"[警告] {name}: count={count}, grids={grids}, avg={avg} 三者不自洽 "
                  f"(实际 avg={grids / count:.4f}, 容差 ±{error})")
            return []
        return [(count, grids)]

    if count is not None and grids is not None:
        if count <= 0 or grids < count:
            print(f"[警告] {name}: count={count}, grids={grids} 违反物理约束")
            return []
        return [(count, grids)]

    if count is not None and avg is not None:
        avg_f = float(avg)
        if count <= 0 or avg_f <= 0:
            return []
        g_lo = max(count, math.ceil((avg_f - error) * count - 1e-9))
        g_hi = min(max_grids, math.floor((avg_f + error) * count + 1e-9))
        return [(count, g) for g in range(g_lo, g_hi + 1)]

    if grids is not None and avg is not None:
        avg_f = float(avg)
        if grids <= 0 or avg_f <= 0:
            return []
        cs = []
        c_hi = min(max_count if max_count else grids, grids)
        for c in range(1, c_hi + 1):
            if abs(grids / c - avg_f) <= error + 1e-9:
                cs.append((c, grids))
        return cs

    if avg is not None:
        return estimate_grid_candidates(avg, max_count or max_grids, max_grids, error)

    if count is not None:
        return [(count, g) for g in range(count, max_grids + 1)]

    if grids is not None:
        c_hi = min(max_count if max_count else grids, grids)
        return [(c, grids) for c in range(1, c_hi + 1)]

    return None


def _apply_total_avg(total_items, total_grids, total_avg, error=0.05):
    if total_avg is None:
        return total_items, total_grids
    avg_f = float(total_avg)
    if total_items is not None and total_grids is None:
        lo = math.ceil((avg_f - error) * total_items - 1e-9)
        hi = math.floor((avg_f + error) * total_items + 1e-9)
        print(f"[信息] 由 total_avg={avg_f} × total_items={total_items} 估算 "
              f"total_grids ∈ [{lo}, {hi}] (作参考, 不锁定)")
        return total_items, None
    if total_grids is not None and total_items is None:
        lo = math.ceil(total_grids / (avg_f + error) - 1e-9)
        hi = math.floor(total_grids / max(avg_f - error, 1e-9) + 1e-9)
        print(f"[信息] 由 total_avg={avg_f} 与 total_grids={total_grids} 估算 "
              f"total_items ∈ [{lo}, {hi}] (作参考, 不锁定)")
        return None, total_grids
    if total_items is not None and total_grids is not None:
        actual = total_grids / total_items
        if not (abs(actual - avg_f) <= error + 1e-9):
            print(f"[警告] total_avg={avg_f} 与 total_items={total_items}, "
                  f"total_grids={total_grids} 不自洽 (实际 avg={actual:.4f})")
    return total_items, total_grids


def combination_analysis(
        total_items=None, total_grids=None, total_avg=None,
        red_value=0, orange_value=0, purple_value=0,
        blue_value=0, green_value=0, white_value=0,
        red_avg=None, orange_avg=None, purple_avg=None,
        blue_avg=None, green_avg=None, white_avg=None,
        red_total_grids=None, orange_total_grids=None, purple_total_grids=None,
        blue_total_grids=None, green_total_grids=None, white_total_grids=None,
        red_count=None, orange_count=None, purple_count=None,
        blue_count=None, green_count=None, white_count=None,
        green_white_total_grids=None, green_white_count=None,
        green_white_avg=None, green_white_value_per_grid=None,
        error=0.05, max_grids_per_quality=200, top_n=10, show_candidates=True,
):
    total_items, total_grids = _apply_total_avg(
        total_items, total_grids, total_avg, error
    )

    green_has_own = (green_avg is not None or green_count is not None
                     or green_total_grids is not None)
    white_has_own = (white_avg is not None or white_count is not None
                     or white_total_grids is not None)

    def _precompute_grids(avg, count, grids, error):
        if grids is not None:
            return grids
        if count is not None and avg is not None:
            cs = _triangulate("_pre", avg, count, None, error,
                              max_count=count, max_grids=max_grids_per_quality)
            if cs and len(cs) == 1:
                return cs[0][1]
        return None

    g_pre = _precompute_grids(green_avg, green_count, green_total_grids, error)
    w_pre = _precompute_grids(white_avg, white_count, white_total_grids, error)
    if green_total_grids is None and g_pre is not None:
        green_total_grids = g_pre
    if white_total_grids is None and w_pre is not None:
        white_total_grids = w_pre

    if green_white_total_grids is not None:
        if green_total_grids is not None and white_total_grids is None:
            white_total_grids = green_white_total_grids - green_total_grids
            print(f"[信息] 由 green_white_total_grids - green_total_grids 推出 "
                  f"white_total_grids = {white_total_grids}")
        elif white_total_grids is not None and green_total_grids is None:
            green_total_grids = green_white_total_grids - white_total_grids
            print(f"[信息] 由 green_white_total_grids - white_total_grids 推出 "
                  f"green_total_grids = {green_total_grids}")

    if green_white_count is not None:
        if green_count is not None and white_count is None:
            white_count = green_white_count - green_count
            print(f"[信息] 由 green_white_count - green_count 推出 "
                  f"white_count = {white_count}")
        elif white_count is not None and green_count is None:
            green_count = green_white_count - white_count
            print(f"[信息] 由 green_white_count - white_count 推出 "
                  f"green_count = {green_count}")

    green_has_own = (green_avg is not None or green_count is not None
                     or green_total_grids is not None)
    white_has_own = (white_avg is not None or white_count is not None
                     or white_total_grids is not None)

    use_green_white_merged = (not green_has_own and not white_has_own
                              and (green_white_total_grids is not None
                                   or green_white_count is not None
                                   or green_white_avg is not None))

    quals = []

    def add_scalar(name, value, avg, count, grids, is_red=False):
        max_count_for_q = total_items if total_items is not None else max_grids_per_quality
        cands = _triangulate(name, avg, count, grids, error,
                             max_count=max_count_for_q, max_grids=max_grids_per_quality)
        if cands is None:
            return
        quals.append({
            "name": name, "kind": "scalar", "value": value,
            "avg": avg, "count": count, "grids": grids,
            "candidates": cands, "is_red": is_red,
        })

    def add_range(name, value_lo, value_hi, value_mid, avg, count, grids):
        max_count_for_q = total_items if total_items is not None else max_grids_per_quality
        cands = _triangulate(name, avg, count, grids, error,
                             max_count=max_count_for_q, max_grids=max_grids_per_quality)
        if cands is None:
            return
        quals.append({
            "name": name, "kind": "range",
            "value_lo": value_lo, "value_hi": value_hi, "value_mid": value_mid,
            "avg": avg, "count": count, "grids": grids,
            "candidates": cands, "is_red": False,
        })

    add_scalar("橙色", orange_value, orange_avg, orange_count, orange_total_grids)
    add_scalar("紫色", purple_value, purple_avg, purple_count, purple_total_grids)
    add_scalar("蓝色", blue_value, blue_avg, blue_count, blue_total_grids)

    if use_green_white_merged:
        lo = min(green_value, white_value)
        hi = max(green_value, white_value)
        mid = green_white_value_per_grid if green_white_value_per_grid is not None \
            else (green_value + white_value) / 2
        add_range("绿白", lo, hi, mid,
                  green_white_avg, green_white_count, green_white_total_grids)
    else:
        add_scalar("绿色", green_value, green_avg, green_count, green_total_grids)
        add_scalar("白色", white_value, white_avg, white_count, white_total_grids)

    red_given = (red_avg is not None or red_count is not None or red_total_grids is not None)
    red_as_derived = not red_given

    if red_given:
        add_scalar("红色", red_value, red_avg, red_count, red_total_grids, is_red=True)

    print(f"========== 组合分析 (容差 ±{error}) ==========")
    if total_items is not None:
        print(f"总件数: {total_items}", end="  ")
    if total_grids is not None:
        print(f"总格数: {total_grids}", end="  ")
    if total_avg is not None:
        print(f"总平均: {total_avg}", end="")
    print()
    if use_green_white_merged:
        print("[模式] 绿白合并 (绿/白无独立信息)")
    else:
        print("[模式] 绿白拆分")
    if red_as_derived:
        print("[红色] 将由 总件数/总格数 差值反推")
    print()

    empty = [q["name"] for q in quals if not q["candidates"]]
    if empty:
        print(f"[错误] 以下品质给定条件下无解: {empty}")
        return []

    if total_items is not None:
        def min_count_of(q):
            vals = [c for c, _ in q["candidates"] if c is not None]
            return min(vals) if vals else 0

        q_min_counts = [min_count_of(q) for q in quals]
        red_min = 0
        total_other_min = sum(q_min_counts) + red_min
        for i, q in enumerate(quals):
            others_min = total_other_min - q_min_counts[i]
            max_c = total_items - others_min
            before = len(q["candidates"])
            q["candidates"] = [(c, g) for c, g in q["candidates"]
                               if c is None or c <= max_c]
            q["prune_count"] = (before, len(q["candidates"]), max_c)

    if total_grids is not None:
        def min_grid_of(q):
            vals = [g for _, g in q["candidates"] if g is not None]
            return min(vals) if vals else 0

        q_min_grids = [min_grid_of(q) for q in quals]
        total_other_min_g = sum(q_min_grids)
        for i, q in enumerate(quals):
            others_min_g = total_other_min_g - q_min_grids[i]
            max_g = total_grids - others_min_g
            before = len(q["candidates"])
            q["candidates"] = [(c, g) for c, g in q["candidates"]
                               if g is None or g <= max_g]
            q["prune_grid"] = (before, len(q["candidates"]), max_g)

    print("各品质输入与候选:")
    for q in quals:
        v_desc = f"{q['value']:,}/格" if q["kind"] == "scalar" \
            else f"{q['value_lo']:,}~{q['value_hi']:,}/格 (期望 {q['value_mid']:,.0f})"
        info_parts = []
        if q["avg"] is not None: info_parts.append(f"avg={q['avg']}")
        if q["count"] is not None: info_parts.append(f"count={q['count']}")
        if q["grids"] is not None: info_parts.append(f"grids={q['grids']}")
        info = ", ".join(info_parts) if info_parts else "(无输入)"

        prune_info = ""
        if "prune_count" in q:
            b, a, mc = q["prune_count"]
            prune_info += f"  件数剪枝 {b}→{a}(上限{mc})"
        if "prune_grid" in q:
            b, a, mg = q["prune_grid"]
            prune_info += f"  格数剪枝 {b}→{a}(上限{mg})"
        print(f"  {q['name']:<3}  [{info}]  {v_desc}  候选 {len(q['candidates'])}{prune_info}")

    if show_candidates:
        for q in quals:
            label_parts = []
            if q["avg"] is not None: label_parts.append(f"avg={q['avg']}")
            if q["count"] is not None: label_parts.append(f"count={q['count']}")
            if q["grids"] is not None: label_parts.append(f"grids={q['grids']}")
            print_candidates_table(q["name"], ",".join(label_parts) or "无输入",
                                   q["candidates"], max_show=40)

    empty2 = [q["name"] for q in quals if not q["candidates"]]
    if empty2:
        print(f"\n[警告] 剪枝后无解: {empty2}")
        return []

    cand_lists = [q["candidates"] for q in quals]
    valid_combos = []

    for combo in product(*cand_lists):
        counts = [c for c, _ in combo]
        grids = [g for _, g in combo]

        if any(c is None for c in counts) or any(g is None for g in grids):
            continue

        sum_count = sum(counts)
        sum_grids = sum(grids)

        if red_as_derived:
            if total_items is None and total_grids is None:
                red_count_v = "?"
                red_grid_v = None
            elif total_items is not None:
                red_count_v = total_items - sum_count
                if red_count_v < 0:
                    continue
                if total_grids is not None:
                    red_grid_v = total_grids - sum_grids
                    if red_grid_v < 0: continue
                    if red_count_v > red_grid_v: continue
                    if red_count_v == 0 and red_grid_v != 0: continue
                else:
                    red_grid_v = None
                    if red_count_v > 0 and red_count_v > max_grids_per_quality:
                        continue
            else:
                red_grid_v = total_grids - sum_grids
                if red_grid_v < 0: continue
                red_count_v = "?"
        else:
            red_count_v = 0
            red_grid_v = 0
            if total_items is not None and sum_count != total_items:
                continue
            if total_grids is not None and sum_grids != total_grids:
                continue

        if total_items is not None and isinstance(red_count_v, int):
            if sum_count + red_count_v != total_items:
                continue
        if total_grids is not None and isinstance(red_grid_v, int):
            if sum_grids + red_grid_v != total_grids:
                continue

        v_lo = v_hi = v_mid = 0
        detail = []
        for i, q in enumerate(quals):
            c, g = combo[i]
            if q["kind"] == "scalar":
                v = g * q["value"]
                v_lo += v; v_hi += v; v_mid += v
            else:
                v_lo += g * q["value_lo"]
                v_hi += g * q["value_hi"]
                v_mid += g * q["value_mid"]
            detail.append((q["name"], c, g))

        if red_as_derived:
            if isinstance(red_grid_v, int):
                rv = red_grid_v * red_value
                v_lo += rv; v_hi += rv; v_mid += rv
            detail.append(("红色", red_count_v,
                           red_grid_v if red_grid_v is not None else "?"))

        valid_combos.append({
            "detail": detail, "v_lo": v_lo, "v_hi": v_hi, "v_mid": v_mid,
        })

    print(f"\n有效组合数: {len(valid_combos)}")
    if not valid_combos:
        print("没有找到满足所有约束的组合")
        return []

    v_mids = sorted(c["v_mid"] for c in valid_combos)
    print(f"\n总价值分布 (按期望值):")
    print(f"  最小期望: {v_mids[0]:>15,.0f}")
    print(f"  最大期望: {v_mids[-1]:>15,.0f}")
    print(f"  平均期望: {sum(v_mids) / len(v_mids):>15,.0f}")
    print(f"  中位数:   {v_mids[len(v_mids) // 2]:>15,.0f}")
    if any(c["v_lo"] != c["v_hi"] for c in valid_combos):
        print(f"\n总价值区间 (考虑绿/白合并不确定):")
        print(f"  整体下界: {min(c['v_lo'] for c in valid_combos):>15,.0f}")
        print(f"  整体上界: {max(c['v_hi'] for c in valid_combos):>15,.0f}")

    def fmt_line(c):
        det = ", ".join(f"{n}({cc}件,{gg}格)" for n, cc, gg in c["detail"])
        if c["v_lo"] == c["v_hi"]:
            return f"  价值 {c['v_mid']:>13,.0f}   |   {det}"
        return (f"  [{c['v_lo']:>13,.0f} ~ {c['v_hi']:>13,.0f}] 期望 "
                f"{c['v_mid']:>13,.0f}   |   {det}")

    valid_combos.sort(key=lambda x: x["v_mid"])
    n_show = min(top_n, len(valid_combos))
    print(f"\n--- 期望总价值 最低 {n_show} 组 ---")
    for c in valid_combos[:n_show]:
        print(fmt_line(c))
    if len(valid_combos) > n_show:
        print(f"\n--- 期望总价值 最高 {n_show} 组 ---")
        for c in valid_combos[-n_show:]:
            print(fmt_line(c))

    return valid_combos


# ============================================================
# 后台计算线程
# ============================================================

class AnalysisWorker(QThread):
    """在后台线程中执行计算，避免界面冻结"""
    finished = pyqtSignal(str, list)  # (输出文本, 组合结果列表)
    error_occurred = pyqtSignal(str)

    def __init__(self, params):
        super().__init__()
        self.params = params

    def run(self):
        try:
            buf = io.StringIO()
            with redirect_stdout(buf):
                results = combination_analysis(**self.params)
            self.finished.emit(buf.getvalue(), results)
        except Exception as e:
            self.error_occurred.emit(str(e))


# ============================================================
# 可选输入框组件 (支持留空 = None)
# ============================================================

class OptionalSpinBox(QSpinBox):
    """支持留空的整数输入框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimum(0)
        self.setMaximum(999999)
        self.setSpecialValueText("未知")
        self.setValue(0)

    def get_value(self):
        val = self.value()
        return None if val == 0 else val

    def reset_to_unknown(self):
        """重置为未知状态"""
        self.setValue(0)


class OptionalDoubleSpinBox(QDoubleSpinBox):
    """支持留空的浮点数输入框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimum(0.0)
        self.setMaximum(999999.0)
        self.setDecimals(2)
        self.setSpecialValueText("未知")
        self.setValue(0.0)

    def get_value(self):
        val = self.value()
        return None if val == 0.0 else val

    def reset_to_unknown(self):
        """重置为未知状态"""
        self.setValue(0.0)


# ============================================================
# 主窗口
# ============================================================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("藏品价值分析工具")
        self.setGeometry(100, 100, 1400, 900)

        # 数据存储
        self.grid_values = {
            "红色": 80000, "橙色": 13000, "紫色": 2400,
            "蓝色": 850, "绿色": 350, "白色": 150
        }
        self.worker = None

        self.init_ui()
        self.apply_styles()

    def init_ui(self):
        """初始化界面"""
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)

        # 主分割器：左侧输入，右侧输出
        splitter = QSplitter(Qt.Horizontal)

        # 左侧：输入区域
        left_widget = self.create_input_area()
        splitter.addWidget(left_widget)

        # 右侧：输出区域
        right_widget = self.create_output_area()
        splitter.addWidget(right_widget)

        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 6)
        main_layout.addWidget(splitter)

        # 状态栏
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage("就绪")

    def create_input_area(self):
        """创建左侧输入区域"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # 使用滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)

        # 全场参数
        global_group = self.create_global_params_group()
        scroll_layout.addWidget(global_group)

        # 每格价值
        value_group = self.create_value_group()
        scroll_layout.addWidget(value_group)

        # 品质参数（使用标签页）
        quality_tabs = self.create_quality_tabs()
        scroll_layout.addWidget(quality_tabs)

        # 绿白合并参数
        green_white_group = self.create_green_white_group()
        scroll_layout.addWidget(green_white_group)

        # 高级参数
        advanced_group = self.create_advanced_group()
        scroll_layout.addWidget(advanced_group)

        # 底部按钮
        button_layout = QHBoxLayout()

        calc_btn = QPushButton("开始计算")
        calc_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px;")
        calc_btn.clicked.connect(self.run_analysis)
        button_layout.addWidget(calc_btn)

        clear_btn = QPushButton("清空参数")
        clear_btn.setStyleSheet("padding: 10px;")
        clear_btn.clicked.connect(self.clear_params)
        button_layout.addWidget(clear_btn)

        scroll_layout.addLayout(button_layout)
        scroll_layout.addStretch()

        scroll.setWidget(scroll_content)
        layout.addWidget(scroll)

        return widget

    def create_global_params_group(self):
        """全场参数组"""
        group = QGroupBox("全场参数")
        layout = QGridLayout()

        self.total_items_spin = OptionalSpinBox()
        self.total_grids_spin = OptionalSpinBox()
        self.total_avg_spin = OptionalDoubleSpinBox()

        # 设置标签字体大小
        label_font = QFont()
        label_font.setPointSize(11)

        label1 = QLabel("总件数:")
        label1.setFont(label_font)
        label2 = QLabel("总格数:")
        label2.setFont(label_font)
        label3 = QLabel("总平均格数:")
        label3.setFont(label_font)

        layout.addWidget(label1, 0, 0)
        layout.addWidget(self.total_items_spin, 0, 1)
        layout.addWidget(label2, 1, 0)
        layout.addWidget(self.total_grids_spin, 1, 1)
        layout.addWidget(label3, 2, 0)
        layout.addWidget(self.total_avg_spin, 2, 1)

        group.setLayout(layout)
        return group

    def create_value_group(self):
        """每格价值组"""
        group = QGroupBox("每格价值")
        layout = QGridLayout()

        self.value_spins = {}
        qualities = ["红色", "橙色", "紫色", "蓝色", "绿色", "白色"]
        colors = ["#FF4444", "#FF8800", "#AA00FF", "#0088FF", "#00CC00", "#888888"]

        # 设置标签字体
        label_font = QFont()
        label_font.setPointSize(11)

        for i, (q, color) in enumerate(zip(qualities, colors)):
            label = QLabel(f"{q}:")
            label.setStyleSheet(f"color: {color}; font-weight: bold;")
            label.setFont(label_font)

            spin = QSpinBox()
            spin.setMinimum(0)
            spin.setMaximum(999999)
            spin.setValue(self.grid_values[q])
            self.value_spins[q] = spin

            row = i // 2
            col = (i % 2) * 2
            layout.addWidget(label, row, col)
            layout.addWidget(spin, row, col + 1)

        group.setLayout(layout)
        return group

    def create_quality_tabs(self):
        """品质参数标签页"""
        tabs = QTabWidget()
        qualities = ["红色", "橙色", "紫色", "蓝色", "绿色", "白色"]
        colors = ["#FF4444", "#FF8800", "#AA00FF", "#0088FF", "#00CC00", "#888888"]

        self.quality_inputs = {}

        # 设置标签字体
        label_font = QFont()
        label_font.setPointSize(11)

        for q, color in zip(qualities, colors):
            tab = QWidget()
            layout = QGridLayout(tab)

            avg_spin = OptionalDoubleSpinBox()
            count_spin = OptionalSpinBox()
            grids_spin = OptionalSpinBox()

            label1 = QLabel("平均格数:")
            label1.setFont(label_font)
            label2 = QLabel("件数:")
            label2.setFont(label_font)
            label3 = QLabel("总格数:")
            label3.setFont(label_font)

            layout.addWidget(label1, 0, 0)
            layout.addWidget(avg_spin, 0, 1)
            layout.addWidget(label2, 1, 0)
            layout.addWidget(count_spin, 1, 1)
            layout.addWidget(label3, 2, 0)
            layout.addWidget(grids_spin, 2, 1)
            layout.setRowStretch(3, 1)

            self.quality_inputs[q] = {
                "avg": avg_spin,
                "count": count_spin,
                "grids": grids_spin
            }

            tabs.addTab(tab, q)
            tabs.tabBar().setTabTextColor(tabs.count() - 1, QColor(color))

        return tabs

    def create_green_white_group(self):
        """绿白合并参数组"""
        group = QGroupBox("绿白合并参数（仅在绿/白无独立信息时使用）")
        layout = QGridLayout()

        self.gw_total_grids_spin = OptionalSpinBox()
        self.gw_count_spin = OptionalSpinBox()
        self.gw_avg_spin = OptionalDoubleSpinBox()
        self.gw_value_spin = OptionalSpinBox()

        # 设置标签字体
        label_font = QFont()
        label_font.setPointSize(11)

        label1 = QLabel("绿白总格数:")
        label1.setFont(label_font)
        label2 = QLabel("绿白总件数:")
        label2.setFont(label_font)
        label3 = QLabel("绿白平均格数:")
        label3.setFont(label_font)
        label4 = QLabel("绿白每格价值:")
        label4.setFont(label_font)

        layout.addWidget(label1, 0, 0)
        layout.addWidget(self.gw_total_grids_spin, 0, 1)
        layout.addWidget(label2, 1, 0)
        layout.addWidget(self.gw_count_spin, 1, 1)
        layout.addWidget(label3, 2, 0)
        layout.addWidget(self.gw_avg_spin, 2, 1)
        layout.addWidget(label4, 3, 0)
        layout.addWidget(self.gw_value_spin, 3, 1)

        group.setLayout(layout)
        return group

    def create_advanced_group(self):
        """高级参数组"""
        group = QGroupBox("高级参数")
        layout = QGridLayout()

        self.error_spin = QDoubleSpinBox()
        self.error_spin.setMinimum(0.0)
        self.error_spin.setMaximum(1.0)
        self.error_spin.setSingleStep(0.01)
        self.error_spin.setValue(0.05)
        self.error_spin.setDecimals(3)

        self.max_grids_spin = QSpinBox()
        self.max_grids_spin.setMinimum(1)
        self.max_grids_spin.setMaximum(300)
        self.max_grids_spin.setValue(80)

        self.top_n_spin = QSpinBox()
        self.top_n_spin.setMinimum(1)
        self.top_n_spin.setMaximum(100)
        self.top_n_spin.setValue(6)

        self.show_candidates_check = QCheckBox()
        self.show_candidates_check.setChecked(True)

        # 设置标签字体
        label_font = QFont()
        label_font.setPointSize(11)

        label1 = QLabel("容差 (±):")
        label1.setFont(label_font)
        label2 = QLabel("单品质最大格数:")
        label2.setFont(label_font)
        label3 = QLabel("显示组合数:")
        label3.setFont(label_font)
        label4 = QLabel("显示候选详情:")
        label4.setFont(label_font)

        layout.addWidget(label1, 0, 0)
        layout.addWidget(self.error_spin, 0, 1)
        layout.addWidget(label2, 1, 0)
        layout.addWidget(self.max_grids_spin, 1, 1)
        layout.addWidget(label3, 2, 0)
        layout.addWidget(self.top_n_spin, 2, 1)
        layout.addWidget(label4, 3, 0)
        layout.addWidget(self.show_candidates_check, 3, 1)

        group.setLayout(layout)
        return group

    def create_output_area(self):
        """创建右侧输出区域"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # 输出文本框
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        self.output_text.setFont(QFont("Consolas", 9))
        layout.addWidget(QLabel("计算输出:"))
        layout.addWidget(self.output_text)

        return widget

    def apply_styles(self):
        """应用样式"""
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #CCCCCC;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton {
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #E0E0E0;
            }
            QLineEdit, QSpinBox, QDoubleSpinBox {
                padding: 3px;
                border: 1px solid #CCCCCC;
                border-radius: 3px;
            }
        """)

    def collect_params(self):
        """收集所有输入参数"""
        params = {
            # 全场
            "total_items": self.total_items_spin.get_value(),
            "total_grids": self.total_grids_spin.get_value(),
            "total_avg": self.total_avg_spin.get_value(),

            # 每格价值
            "red_value": self.value_spins["红色"].value(),
            "orange_value": self.value_spins["橙色"].value(),
            "purple_value": self.value_spins["紫色"].value(),
            "blue_value": self.value_spins["蓝色"].value(),
            "green_value": self.value_spins["绿色"].value(),
            "white_value": self.value_spins["白色"].value(),

            # 绿白合并
            "green_white_total_grids": self.gw_total_grids_spin.get_value(),
            "green_white_count": self.gw_count_spin.get_value(),
            "green_white_avg": self.gw_avg_spin.get_value(),
            "green_white_value_per_grid": self.gw_value_spin.get_value(),

            # 高级
            "error": self.error_spin.value(),
            "max_grids_per_quality": self.max_grids_spin.value(),
            "top_n": self.top_n_spin.value(),
            "show_candidates": self.show_candidates_check.isChecked(),
        }

        # 各品质参数
        quality_map = {
            "红色": "red", "橙色": "orange", "紫色": "purple",
            "蓝色": "blue", "绿色": "green", "白色": "white"
        }
        for q_cn, q_en in quality_map.items():
            inputs = self.quality_inputs[q_cn]
            params[f"{q_en}_avg"] = inputs["avg"].get_value()
            params[f"{q_en}_count"] = inputs["count"].get_value()
            params[f"{q_en}_total_grids"] = inputs["grids"].get_value()

        return params

    def run_analysis(self):
        """运行分析"""
        if self.worker and self.worker.isRunning():
            QMessageBox.warning(self, "警告", "计算正在进行中，请稍候...")
            return

        params = self.collect_params()

        self.output_text.clear()
        self.statusBar.showMessage("正在计算...")

        self.worker = AnalysisWorker(params)
        self.worker.finished.connect(self.on_analysis_finished)
        self.worker.error_occurred.connect(self.on_analysis_error)
        self.worker.start()

    def on_analysis_finished(self, output, results):
        """计算完成"""
        self.output_text.setPlainText(output)
        self.statusBar.showMessage(f"计算完成，找到 {len(results)} 个有效组合", 5000)

        # 滚动到底部
        cursor = self.output_text.textCursor()
        cursor.movePosition(QTextCursor.End)
        self.output_text.setTextCursor(cursor)

    def on_analysis_error(self, error_msg):
        """计算出错"""
        self.statusBar.showMessage("计算失败", 3000)
        self.output_text.append(f"\n[错误] {error_msg}")
        QMessageBox.critical(self, "计算失败", f"计算过程中出现错误:\n{error_msg}")

    def clear_params(self):
        """清空参数（保留每格价值和高级参数）"""
        reply = QMessageBox.question(
            self, "确认清空",
            "确定要清空参数吗？\n（每格价值和高级参数将保留）",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # 清空全场参数
            self.total_items_spin.reset_to_unknown()
            self.total_grids_spin.reset_to_unknown()
            self.total_avg_spin.reset_to_unknown()

            # 清空各品质参数
            for q in ["红色", "橙色", "紫色", "蓝色", "绿色", "白色"]:
                self.quality_inputs[q]["avg"].reset_to_unknown()
                self.quality_inputs[q]["count"].reset_to_unknown()
                self.quality_inputs[q]["grids"].reset_to_unknown()

            # 清空绿白合并参数
            self.gw_total_grids_spin.reset_to_unknown()
            self.gw_count_spin.reset_to_unknown()
            self.gw_avg_spin.reset_to_unknown()
            self.gw_value_spin.reset_to_unknown()

            # 清空输出
            self.output_text.clear()

            self.statusBar.showMessage("参数已清空（每格价值和高级参数已保留）", 3000)


# ============================================================
# 程序入口
# ============================================================

def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # 使用 Fusion 风格获得更好的跨平台外观

    window = MainWindow()
    window.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()