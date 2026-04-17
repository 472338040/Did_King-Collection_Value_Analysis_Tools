import os
import math
from math import gcd
from itertools import product

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# ============================================================
# 从 Data.xlsx 读取每个品质的统计数据
# ============================================================
def load_grid_stats(xlsx_path="Data.xlsx", weighted=False, verbose=True):
    """
    读取 Excel 中的物品均价数据，对每个品质统计:
        - per_grid : 平均每格价值
        - avg_grids: 每件物品平均占用格数
        - samples  : 样本数
    另外合成 "绿白" 条目 (绿色+白色 样本并在一起).
    """
    if load_workbook is None:
        raise RuntimeError("未安装 openpyxl，请先执行: pip install openpyxl")
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"找不到文件: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    col_to_grids = {}
    for cell in ws[1]:
        val = cell.value
        if isinstance(val, str) and "*" in val:
            left = val.split("*", 1)[0].strip()
            try:
                col_to_grids[cell.column] = int(left)
            except ValueError:
                continue
    if not col_to_grids:
        raise ValueError("表头解析失败，未找到形如 'X*Y物品均价' 的列")

    raw_samples = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        quality = row[0].value
        if not quality:
            continue
        samples = []
        for cell in row[1:]:
            grids = col_to_grids.get(cell.column)
            if grids is None:
                continue
            price = cell.value
            if not isinstance(price, (int, float)) or price <= 0:
                continue
            samples.append((price, grids))
        raw_samples[quality] = samples

    raw_samples["绿白"] = raw_samples.get("绿色", []) + raw_samples.get("白色", [])

    results = {}
    if verbose:
        mode_name = "加权平均" if weighted else "简单平均"
        print(f"========== 从 {xlsx_path} 读取统计 ({mode_name}) ==========")
        print(f"  {'品质':<4}  {'样本数':>5}  {'每格价值':>12}  {'平均格数':>10}")
        print("  " + "-" * 40)

    for quality, samples in raw_samples.items():
        if not samples:
            results[quality] = {"per_grid": 0, "avg_grids": 0.0, "samples": 0}
            if verbose:
                print(f"  [警告] {quality}: 没有有效数据")
            continue
        per_grid_prices = [p / g for p, g in samples]
        total_price = sum(p for p, _ in samples)
        total_grids_sum = sum(g for _, g in samples)
        avg_price_per_grid = (total_price / total_grids_sum) if weighted \
            else sum(per_grid_prices) / len(per_grid_prices)
        avg_grids_per_item = total_grids_sum / len(samples)
        results[quality] = {
            "per_grid": int(round(avg_price_per_grid)),
            "avg_grids": round(avg_grids_per_item, 4),
            "samples": len(samples),
        }
        if verbose:
            print(f"  {quality:<4}  {len(samples):>5d}  "
                  f"{results[quality]['per_grid']:>12,}  "
                  f"{results[quality]['avg_grids']:>10.3f}")

    if verbose:
        print("=" * 56)
        print()
    return results


def load_grid_values(xlsx_path="Data.xlsx", weighted=False, verbose=True):
    """向后兼容: 仅返回 {品质: 每格价值}。"""
    stats = load_grid_stats(xlsx_path, weighted, verbose)
    return {q: s["per_grid"] for q, s in stats.items()}


# ============================================================
# 核心: 根据平均格数 + 容差, 枚举所有合法的 (件数, 总格数) 候选
# ============================================================
def estimate_grid_candidates(avg, max_count, max_grids=200, error=0.05,
                             known_grids=None):
    """
    给定显示的平均格数 avg, 找到所有满足
        |grids / count - avg| <= error
    的 (count, grids) 组合.

    参数:
        avg         : 显示的平均格数 (float/str). 若同时有 known_grids 可为 None.
        max_count   : 件数上限 (通常是全场总藏品数)
        max_grids   : 单品质总格数上限 (默认 200); 有 known_grids 时忽略
        error       : 容差 (默认 ±0.05)
        known_grids : 已知该品质总格数. 若给出:
                       - 有 avg: 只保留 g == known_grids 且满足容差的 count
                       - 无 avg: 返回 [(c, known_grids) for c in 1..min(max_count, known_grids)]
                                (件数完全未知, 组合时再被其他约束收缩)

    返回:
        [(count, grids), ...]  按 count 升序
    """
    # 情况 A: 已知总格数
    if known_grids is not None:
        g = known_grids
        if avg is None:
            # 完全没有 avg 约束, 件数只能被物理约束 1 <= c <= min(max_count, g)
            # (每件至少占 1 格)
            return [(c, g) for c in range(1, min(max_count, g) + 1)]
        avg_f = float(avg)
        if avg_f <= 0:
            return []
        candidates = []
        for count in range(1, min(max_count, g) + 1):
            if abs(g / count - avg_f) <= error + 1e-9:
                candidates.append((count, g))
        return candidates

    # 情况 B: 仅知 avg
    if avg is None:
        return []
    avg = float(avg)
    if avg <= 0:
        return []

    candidates = []
    for count in range(1, max_count + 1):
        min_g = (avg - error) * count
        max_g = (avg + error) * count
        g_low = max(0, math.ceil(min_g - 1e-9))
        g_high = min(max_grids, math.floor(max_g + 1e-9))
        for g in range(g_low, g_high + 1):
            if count == 0:
                continue
            if g < count:          # 每件至少 1 格的物理约束
                continue
            if abs(g / count - avg) <= error + 1e-9:
                candidates.append((count, g))
    return candidates


def print_candidates_table(name, label, candidates, max_show=50):
    """打印候选表, 方便人工查看."""
    print(f"\n[{name}]  {label}  候选 {len(candidates)} 个")
    if not candidates:
        print("  (无)")
        return
    shown = candidates[:max_show]
    line = "  ".join(f"({c},{g})" for c, g in shown)
    print(f"  (count, grids): {line}"
          + (f"  ... 还有 {len(candidates) - max_show} 个" if len(candidates) > max_show else ""))


# ============================================================
# 核心: 跨品质枚举组合, 保留件数合计 <= 总藏品数, 并计算总价值
# ============================================================

# ---------- 单品质三变量互推 (avg, count, total_grids) ----------
def _triangulate(name, avg, count, grids, error=0.05, max_count=None, max_grids=200):
    """
    根据已知的任意 1~3 个槽位, 返回候选 [(count, grids), ...].

    规则:
      - 三个都已知: 先校验一致性 |grids/count - avg| <= error, 通过则返回 [(count, grids)]
      - 已知 count + grids: 直接返回 [(count, grids)] (avg 反推)
      - 已知 count + avg:   枚举 g 使 |g/count - avg|<=error 且 g>=count
      - 已知 grids + avg:   枚举 c 使 |grids/c - avg|<=error 且 c<=grids
      - 已知 avg:           枚举 (c, g) 使 |g/c-avg|<=error 且 c<=g<=max_grids
      - 已知 count:         c 已知, g 未知 → [(c, None)]  (交给后续约束)
      - 已知 grids:         g 已知, c 未知 → [(None, g)]  (交给后续约束)
      - 什么都没知道: 返回 None 表示"该品质不参与分析"
    """
    if count is None and grids is None and avg is None:
        return None

    # ---- 边界: count=0 → grids 必须是 0; grids=0 → count 必须是 0 ----
    if count is not None and count == 0:
        if grids is not None and grids != 0:
            print(f"[警告] {name}: count=0 但 grids={grids}, 矛盾")
            return []
        return [(0, 0)]
    if grids is not None and grids == 0:
        if count is not None and count != 0:
            print(f"[警告] {name}: grids=0 但 count={count}, 矛盾")
            return []
        return [(0, 0)]

    # 三者都给
    if count is not None and grids is not None and avg is not None:
        if count <= 0 or grids < count:
            print(f"[警告] {name}: count={count}, grids={grids} 违反物理约束 count<=grids")
            return []
        if not (abs(grids / count - float(avg)) <= error + 1e-9):
            print(f"[警告] {name}: count={count}, grids={grids}, avg={avg} 三者不自洽 "
                  f"(实际 avg={grids/count:.4f}, 容差 ±{error})")
            return []
        return [(count, grids)]

    # 已知 count + grids (无 avg)
    if count is not None and grids is not None:
        if count <= 0 or grids < count:
            print(f"[警告] {name}: count={count}, grids={grids} 违反物理约束")
            return []
        return [(count, grids)]

    # 已知 count + avg (无 grids)
    if count is not None and avg is not None:
        avg_f = float(avg)
        if count <= 0 or avg_f <= 0:
            return []
        g_lo = max(count, math.ceil((avg_f - error) * count - 1e-9))
        g_hi = min(max_grids, math.floor((avg_f + error) * count + 1e-9))
        return [(count, g) for g in range(g_lo, g_hi + 1)]

    # 已知 grids + avg (无 count)
    if grids is not None and avg is not None:
        avg_f = float(avg)
        if grids <= 0 or avg_f <= 0:
            return []
        cs = []
        c_hi = min(max_count if max_count else grids, grids)
        for c in range(1, c_hi + 1):
            if abs(grids / c - avg_f) <= error + 1e-9:
                cs.append((c, grids))
        return cs

    # 只有 avg
    if avg is not None:
        return estimate_grid_candidates(avg, max_count or max_grids, max_grids, error)

    # 只有 count (无 grids, 无 avg) → 枚举 g ∈ [count, max_grids]
    if count is not None:
        return [(count, g) for g in range(count, max_grids + 1)]

    # 只有 grids (无 count, 无 avg) → 格数锁定, 件数任意
    if grids is not None:
        c_hi = min(max_count if max_count else grids, grids)
        return [(c, grids) for c in range(1, c_hi + 1)]

    return None


def _apply_total_avg(total_items, total_grids, total_avg, error=0.05):
    """
    用 total_avg 互推/校验 total_items 与 total_grids.
    返回 (total_items, total_grids), 不一致则打印警告.
    """
    if total_avg is None:
        return total_items, total_grids
    avg_f = float(total_avg)
    if total_items is not None and total_grids is None:
        # 按 avg 估算 grids 区间, 但不强行锁定, 留 None 让下游用件数推
        lo = math.ceil((avg_f - error) * total_items - 1e-9)
        hi = math.floor((avg_f + error) * total_items + 1e-9)
        print(f"[信息] 由 total_avg={avg_f} × total_items={total_items} 估算 "
              f"total_grids ∈ [{lo}, {hi}] (作参考, 不锁定)")
        return total_items, None
    if total_grids is not None and total_items is None:
        # 同上, 估算 total_items 区间
        lo = math.ceil(total_grids / (avg_f + error) - 1e-9)
        hi = math.floor(total_grids / max(avg_f - error, 1e-9) + 1e-9)
        print(f"[信息] 由 total_avg={avg_f} 与 total_grids={total_grids} 估算 "
              f"total_items ∈ [{lo}, {hi}] (作参考, 不锁定)")
        return None, total_grids
    if total_items is not None and total_grids is not None:
        actual = total_grids / total_items
        if not (abs(actual - avg_f) <= error + 1e-9):
            print(f"[警告] total_avg={avg_f} 与 total_items={total_items}, "
                  f"total_grids={total_grids} 不自洽 (实际 avg={actual:.4f})")
    return total_items, total_grids


# ============================================================
# 核心: 六品质联合估算
# ============================================================
def combination_analysis(
    # ---- 全场参数 ----
    total_items=None,
    total_grids=None,
    total_avg=None,

    # ---- 每格价值 ----
    red_value=0, orange_value=0, purple_value=0,
    blue_value=0, green_value=0, white_value=0,

    # ---- 各品质平均格数 ----
    red_avg=None, orange_avg=None, purple_avg=None,
    blue_avg=None, green_avg=None, white_avg=None,

    # ---- 各品质总格数 ----
    red_total_grids=None, orange_total_grids=None, purple_total_grids=None,
    blue_total_grids=None, green_total_grids=None, white_total_grids=None,

    # ---- 各品质件数 ----
    red_count=None, orange_count=None, purple_count=None,
    blue_count=None, green_count=None, white_count=None,

    # ---- 绿白合并信息 (仅当绿或白缺乏独立信息时使用) ----
    green_white_total_grids=None,
    green_white_count=None,
    green_white_avg=None,
    green_white_value_per_grid=None,

    # ---- 枚举参数 ----
    error=0.05,
    max_grids_per_quality=200,
    top_n=10,
    show_candidates=True,
):
    """
    六品质联合估算.

    对每个品质, 根据已知的 (avg, count, total_grids) 中任意子集生成候选;
    绿色缺乏独立信息但有 green_white_* 时, 退化成"绿白"合并品质.
    再对所有品质做笛卡尔积, 应用全场 total_items / total_grids 约束,
    最后计算总价值分布.
    """
    total_items, total_grids = _apply_total_avg(
        total_items, total_grids, total_avg, error
    )

    # ---------- 绿白拆分/合并 决策 ----------
    green_has_own = (green_avg is not None or green_count is not None
                     or green_total_grids is not None)
    white_has_own = (white_avg is not None or white_count is not None
                     or white_total_grids is not None)

    # 先对 绿/白 做一次独立预推, 尝试把 total_grids 锁死, 给联动用
    def _precompute_grids(avg, count, grids, error):
        """若能唯一锁定 total_grids 就返回它, 否则返回 None."""
        if grids is not None:
            return grids
        if count is not None and avg is not None:
            cs = _triangulate("_pre", avg, count, None, error,
                              max_count=count, max_grids=max_grids_per_quality)
            if cs and len(cs) == 1:
                return cs[0][1]
        return None

    g_pre = _precompute_grids(green_avg, green_count, green_total_grids, error)
    w_pre = _precompute_grids(white_avg, white_count, white_total_grids, error)
    if green_total_grids is None and g_pre is not None:
        green_total_grids = g_pre
    if white_total_grids is None and w_pre is not None:
        white_total_grids = w_pre

    # 先尝试用 green_white_total_grids 补全 white/green 总格数
    if green_white_total_grids is not None:
        if green_total_grids is not None and white_total_grids is None:
            white_total_grids = green_white_total_grids - green_total_grids
            print(f"[信息] 由 green_white_total_grids - green_total_grids 推出 "
                  f"white_total_grids = {white_total_grids}")
        elif white_total_grids is not None and green_total_grids is None:
            green_total_grids = green_white_total_grids - white_total_grids
            print(f"[信息] 由 green_white_total_grids - white_total_grids 推出 "
                  f"green_total_grids = {green_total_grids}")

    # 用 green_white_count 补全 white_count / green_count
    if green_white_count is not None:
        if green_count is not None and white_count is None:
            white_count = green_white_count - green_count
            print(f"[信息] 由 green_white_count - green_count 推出 "
                  f"white_count = {white_count}")
        elif white_count is not None and green_count is None:
            green_count = green_white_count - white_count
            print(f"[信息] 由 green_white_count - white_count 推出 "
                  f"green_count = {green_count}")

    # 重新判断: 补全后是否都有独立信息
    green_has_own = (green_avg is not None or green_count is not None
                     or green_total_grids is not None)
    white_has_own = (white_avg is not None or white_count is not None
                     or white_total_grids is not None)

    use_green_white_merged = (not green_has_own and not white_has_own
                              and (green_white_total_grids is not None
                                   or green_white_count is not None
                                   or green_white_avg is not None))

    # ---------- 构建品质列表 ----------
    quals = []  # 每个元素: {name, kind, value/value_lo/value_hi/value_mid,
                #           avg, count, grids, candidates, is_red}

    def add_scalar(name, value, avg, count, grids, is_red=False):
        max_count_for_q = total_items if total_items is not None else max_grids_per_quality
        cands = _triangulate(name, avg, count, grids, error,
                             max_count=max_count_for_q, max_grids=max_grids_per_quality)
        if cands is None:
            return
        quals.append({
            "name": name, "kind": "scalar", "value": value,
            "avg": avg, "count": count, "grids": grids,
            "candidates": cands, "is_red": is_red,
        })

    def add_range(name, value_lo, value_hi, value_mid, avg, count, grids):
        max_count_for_q = total_items if total_items is not None else max_grids_per_quality
        cands = _triangulate(name, avg, count, grids, error,
                             max_count=max_count_for_q, max_grids=max_grids_per_quality)
        if cands is None:
            return
        quals.append({
            "name": name, "kind": "range",
            "value_lo": value_lo, "value_hi": value_hi, "value_mid": value_mid,
            "avg": avg, "count": count, "grids": grids,
            "candidates": cands, "is_red": False,
        })

    # 红色留到最后单独处理 (由差值反推), 但也可直接作为品质加入
    add_scalar("橙色", orange_value, orange_avg, orange_count, orange_total_grids)
    add_scalar("紫色", purple_value, purple_avg, purple_count, purple_total_grids)
    add_scalar("蓝色", blue_value,   blue_avg,   blue_count,   blue_total_grids)

    if use_green_white_merged:
        lo = min(green_value, white_value)
        hi = max(green_value, white_value)
        mid = green_white_value_per_grid if green_white_value_per_grid is not None \
              else (green_value + white_value) / 2
        add_range("绿白", lo, hi, mid,
                  green_white_avg, green_white_count, green_white_total_grids)
    else:
        add_scalar("绿色", green_value, green_avg, green_count, green_total_grids)
        add_scalar("白色", white_value, white_avg, white_count, white_total_grids)

    # 红色: 如果用户直接给了任何一个槽位, 当作普通品质; 否则后面用差值反推
    red_given = (red_avg is not None or red_count is not None or red_total_grids is not None)
    red_as_derived = not red_given  # 完全没红色信息时, 走"差值反推"分支

    if red_given:
        add_scalar("红色", red_value, red_avg, red_count, red_total_grids, is_red=True)

    # ---------- 打印头部 ----------
    print(f"========== 组合分析 (容差 ±{error}) ==========")
    if total_items is not None:
        print(f"总件数: {total_items}", end="  ")
    if total_grids is not None:
        print(f"总格数: {total_grids}", end="  ")
    if total_avg is not None:
        print(f"总平均: {total_avg}", end="")
    print()
    if use_green_white_merged:
        print("[模式] 绿白合并 (绿/白无独立信息)")
    else:
        print("[模式] 绿白拆分")
    if red_as_derived:
        print("[红色] 将由 总件数/总格数 差值反推")
    print()

    # 检查某个品质候选为空
    empty = [q["name"] for q in quals if not q["candidates"]]
    if empty:
        print(f"[错误] 以下品质给定条件下无解: {empty}")
        return []

    # ---------- 剪枝: 基于件数可行性 ----------
    if total_items is not None:
        # 对每个品质, 其他品质的最小件数合计 + 红色下限
        def min_count_of(q):
            vals = [c for c, _ in q["candidates"] if c is not None]
            return min(vals) if vals else 0
        q_min_counts = [min_count_of(q) for q in quals]
        red_min = 0 if red_as_derived else 0  # red_as_derived 时红色件数 = 剩余件数, 下限 0
        total_other_min = sum(q_min_counts) + red_min
        for i, q in enumerate(quals):
            others_min = total_other_min - q_min_counts[i]
            max_c = total_items - others_min
            before = len(q["candidates"])
            q["candidates"] = [(c, g) for c, g in q["candidates"]
                               if c is None or c <= max_c]
            q["prune_count"] = (before, len(q["candidates"]), max_c)

    # ---------- 剪枝: 基于格数可行性 ----------
    if total_grids is not None:
        def min_grid_of(q):
            vals = [g for _, g in q["candidates"] if g is not None]
            return min(vals) if vals else 0
        q_min_grids = [min_grid_of(q) for q in quals]
        total_other_min_g = sum(q_min_grids)
        for i, q in enumerate(quals):
            others_min_g = total_other_min_g - q_min_grids[i]
            max_g = total_grids - others_min_g
            before = len(q["candidates"])
            q["candidates"] = [(c, g) for c, g in q["candidates"]
                               if g is None or g <= max_g]
            q["prune_grid"] = (before, len(q["candidates"]), max_g)

    # ---------- 打印候选信息 ----------
    print("各品质输入与候选:")
    for q in quals:
        v_desc = f"{q['value']:,}/格" if q["kind"] == "scalar" \
                 else f"{q['value_lo']:,}~{q['value_hi']:,}/格 (期望 {q['value_mid']:,.0f})"
        info_parts = []
        if q["avg"] is not None: info_parts.append(f"avg={q['avg']}")
        if q["count"] is not None: info_parts.append(f"count={q['count']}")
        if q["grids"] is not None: info_parts.append(f"grids={q['grids']}")
        info = ", ".join(info_parts) if info_parts else "(无输入)"

        prune_info = ""
        if "prune_count" in q:
            b, a, mc = q["prune_count"]
            prune_info += f"  件数剪枝 {b}→{a}(上限{mc})"
        if "prune_grid" in q:
            b, a, mg = q["prune_grid"]
            prune_info += f"  格数剪枝 {b}→{a}(上限{mg})"
        print(f"  {q['name']:<3}  [{info}]  {v_desc}  候选 {len(q['candidates'])}{prune_info}")

    if show_candidates:
        for q in quals:
            label_parts = []
            if q["avg"] is not None: label_parts.append(f"avg={q['avg']}")
            if q["count"] is not None: label_parts.append(f"count={q['count']}")
            if q["grids"] is not None: label_parts.append(f"grids={q['grids']}")
            print_candidates_table(q["name"], ",".join(label_parts) or "无输入",
                                   q["candidates"], max_show=40)

    empty2 = [q["name"] for q in quals if not q["candidates"]]
    if empty2:
        print(f"\n[警告] 剪枝后无解: {empty2}")
        return []

    # ---------- 枚举 ----------
    cand_lists = [q["candidates"] for q in quals]
    valid_combos = []

    for combo in product(*cand_lists):
        # 件数合计 (包含已知的红色)
        counts = [c for c, _ in combo]
        grids = [g for _, g in combo]

        # 若组合里有 None, 说明有品质件/格数未锁定 → 跳过 (目前不支持部分未知)
        if any(c is None for c in counts) or any(g is None for g in grids):
            continue

        sum_count = sum(counts)
        sum_grids = sum(grids)

        # 红色处理
        if red_as_derived:
            if total_items is None and total_grids is None:
                # 无法反推红色, 标记未知
                red_count_v = "?"
                red_grid_v = None
            elif total_items is not None:
                red_count_v = total_items - sum_count
                if red_count_v < 0:
                    continue
                if total_grids is not None:
                    red_grid_v = total_grids - sum_grids
                    if red_grid_v < 0: continue
                    if red_count_v > red_grid_v: continue
                    if red_count_v == 0 and red_grid_v != 0: continue
                else:
                    # 只知件数不知格数, 不锁定格数 → 价值部分未知
                    red_grid_v = None
                    if red_count_v > 0 and red_count_v > max_grids_per_quality:
                        continue
            else:
                # 只知 total_grids, 不知 total_items: 红色件数未知, 格数由差值
                red_grid_v = total_grids - sum_grids
                if red_grid_v < 0: continue
                red_count_v = "?"
        else:
            red_count_v = 0    # 所有件数都在 combo 里了
            red_grid_v = 0
            # 校验总件数 / 总格数
            if total_items is not None and sum_count != total_items:
                continue
            if total_grids is not None and sum_grids != total_grids:
                continue

        if total_items is not None and isinstance(red_count_v, int):
            if sum_count + red_count_v != total_items:
                continue
        if total_grids is not None and isinstance(red_grid_v, int):
            if sum_grids + red_grid_v != total_grids:
                continue

        # ---- 价值计算 ----
        v_lo = v_hi = v_mid = 0
        detail = []
        for i, q in enumerate(quals):
            c, g = combo[i]
            if q["kind"] == "scalar":
                v = g * q["value"]
                v_lo += v; v_hi += v; v_mid += v
            else:
                v_lo += g * q["value_lo"]
                v_hi += g * q["value_hi"]
                v_mid += g * q["value_mid"]
            detail.append((q["name"], c, g))

        if red_as_derived:
            if isinstance(red_grid_v, int):
                rv = red_grid_v * red_value
                v_lo += rv; v_hi += rv; v_mid += rv
            detail.append(("红色", red_count_v, red_grid_v if red_grid_v is not None else "?"))

        valid_combos.append({
            "detail": detail, "v_lo": v_lo, "v_hi": v_hi, "v_mid": v_mid,
        })

    # ---------- 输出 ----------
    print(f"\n有效组合数: {len(valid_combos)}")
    if not valid_combos:
        print("没有找到满足所有约束的组合")
        return []

    v_mids = sorted(c["v_mid"] for c in valid_combos)
    print(f"\n总价值分布 (按期望值):")
    print(f"  最小期望: {v_mids[0]:>15,.0f}")
    print(f"  最大期望: {v_mids[-1]:>15,.0f}")
    print(f"  平均期望: {sum(v_mids)/len(v_mids):>15,.0f}")
    print(f"  中位数:   {v_mids[len(v_mids)//2]:>15,.0f}")
    if any(c["v_lo"] != c["v_hi"] for c in valid_combos):
        print(f"\n总价值区间 (考虑绿/白合并不确定):")
        print(f"  整体下界: {min(c['v_lo'] for c in valid_combos):>15,.0f}")
        print(f"  整体上界: {max(c['v_hi'] for c in valid_combos):>15,.0f}")

    def fmt_line(c):
        det = ", ".join(f"{n}({cc}件,{gg}格)" for n, cc, gg in c["detail"])
        if c["v_lo"] == c["v_hi"]:
            return f"  价值 {c['v_mid']:>13,.0f}   |   {det}"
        return (f"  [{c['v_lo']:>13,.0f} ~ {c['v_hi']:>13,.0f}] 期望 "
                f"{c['v_mid']:>13,.0f}   |   {det}")

    valid_combos.sort(key=lambda x: x["v_mid"])
    n_show = min(top_n, len(valid_combos))
    print(f"\n--- 期望总价值 最低 {n_show} 组 ---")
    for c in valid_combos[:n_show]:
        print(fmt_line(c))
    if len(valid_combos) > n_show:
        print(f"\n--- 期望总价值 最高 {n_show} 组 ---")
        for c in valid_combos[-n_show:]:
            print(fmt_line(c))

    return valid_combos


# ============================================================
# 以下为原有分析逻辑 (按最简分数推最小组合), 保留作对照
# ============================================================
def analyze_avg(avg_str, total_items, value_per_grid, known_count=None):
    avg_100 = int(round(float(avg_str) * 100))
    g = gcd(avg_100, 100)
    numerator = avg_100 // g
    denominator = 100 // g
    max_k = total_items // denominator

    if known_count is not None:
        if known_count % denominator == 0:
            total_grids = int(float(avg_str) * known_count)
            total_value = total_grids * value_per_grid
            return {"平均格数": float(avg_str), "最简分数": f"{numerator}/{denominator}",
                    "最小藏品数量": known_count, "最小总格数": total_grids,
                    "第二小藏品数量": "-", "第二小总格数": "-",
                    "单格价值": value_per_grid, "最小总价值": total_value,
                    "第二小总价值": "-", "推算方式": "由平均格数+已知数量推算"}
        return {"平均格数": float(avg_str), "最简分数": f"{numerator}/{denominator}",
                "最小藏品数量": known_count, "最小总格数": "数量不合法",
                "第二小藏品数量": "-", "第二小总格数": "-",
                "单格价值": value_per_grid, "最小总价值": "-",
                "第二小总价值": "-", "推算方式": f"数量需为 {denominator} 的倍数"}

    possible_pairs = [(denominator * k, numerator * k) for k in range(1, max_k + 1)]
    first_pair = possible_pairs[0] if len(possible_pairs) >= 1 else ("-", "-")
    second_pair = possible_pairs[1] if len(possible_pairs) >= 2 else ("-", "-")
    first_value = first_pair[1] * value_per_grid if first_pair[1] != "-" else "-"
    second_value = second_pair[1] * value_per_grid if second_pair[1] != "-" else "-"
    return {"平均格数": float(avg_str), "最简分数": f"{numerator}/{denominator}",
            "最小藏品数量": first_pair[0], "最小总格数": first_pair[1],
            "第二小藏品数量": second_pair[0], "第二小总格数": second_pair[1],
            "单格价值": value_per_grid, "最小总价值": first_value,
            "第二小总价值": second_value, "推算方式": "由平均格数推算"}


# ============================================================
# 示例
# ============================================================
if __name__ == "__main__":
    # ---------- 每格价值 (通常从 Data.xlsx 取) ----------
    # red_value = 165947
    # orange_value = 27416
    # purple_value = 6449
    # blue_value = 2406
    # green_value = 707
    # white_value = 211

    red_value = 165947
    orange_value = 23000
    purple_value = 3000
    blue_value = 900
    green_value = 550
    white_value = 160

    green_white_value_per_grid = None  # 有则更准, 没有用 (green+white)/2

    # ---------- 全场 (回合 0 会给 total_items) ----------
    total_items = 26   # 回合 0 固定
    total_grids = None  # 随机道具才能看到
    total_avg = None  # 随机道具才能看到

    # ---------- 各品质三槽位 (avg / count / total_grids) ----------
    # 没拿到就留 None, 有几条填几条 (能互推的函数内部会自己推)
    red_avg = None
    red_count = None
    red_total_grids = None

    orange_avg = 1  # 回合 1 固定
    orange_count = None
    orange_total_grids = None

    purple_avg = 1.33  # 回合 2 固定
    purple_count = None
    purple_total_grids = 4

    blue_avg = 2.66  # 回合 3 固定
    blue_count = None
    blue_total_grids = None

    green_avg = 1.16  # 回合 2 固定
    green_count = 6  # 回合 1 道具
    green_total_grids = None

    white_avg = None
    white_count = None
    white_total_grids = None

    # ---------- 绿白合并信息 (仅在绿/白都没独立信息时生效) ----------
    # 回合 0 会给 green_white_total_grids; 回合 5 会给 green_white_count
    green_white_total_grids = 17    # 回合 0 道具
    green_white_count = 13          # 回合 5 固定
    green_white_avg = None

    combination_analysis(
        total_items=total_items,
        total_grids=total_grids,
        total_avg=total_avg,

        red_value=red_value, orange_value=orange_value, purple_value=purple_value,
        blue_value=blue_value, green_value=green_value, white_value=white_value,

        red_avg=red_avg, orange_avg=orange_avg, purple_avg=purple_avg,
        blue_avg=blue_avg, green_avg=green_avg, white_avg=white_avg,

        red_count=red_count, orange_count=orange_count, purple_count=purple_count,
        blue_count=blue_count, green_count=green_count, white_count=white_count,

        red_total_grids=red_total_grids, orange_total_grids=orange_total_grids,
        purple_total_grids=purple_total_grids, blue_total_grids=blue_total_grids,
        green_total_grids=green_total_grids, white_total_grids=white_total_grids,

        green_white_total_grids=green_white_total_grids,
        green_white_count=green_white_count,
        green_white_avg=green_white_avg,
        green_white_value_per_grid=green_white_value_per_grid,

        error=0.05,
        top_n=10,
        show_candidates=True,
    )
